# encoding: utf-8
from openpyxl import Workbook
__author__ = 'xunyan'


# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()

# 获得一个sheet页
ws = wb.get_active_sheet()

# 打印sheet页的标题
print(ws.title)

ws.title = 'test'  # 设置worksheet的标题

# 设置单元格的值
ws.cell('D3').value = 4
ws.cell(row=3, column=1).value = 6

# 新打开一个sheet页，sheet页的标题命名为
new_ws = wb.create_sheet(title=u'乘法表')
# row = 1
# col = 2
# new_ws.cell(row=1, column=1).value = str(row) + "*" + str(col) + "=" + str(row*col)

# 向sheet页中写入值
for row in range(1, 10):
    for col in range(1, row+1):
        new_ws.cell(row=row, column=col).value = str(row) + "*" + str(col) + "=" + str(row*col)

# 简单实现方法
# for row in range(1, 10):
#     for col in range(1, 10):
#         sheet1.cell(row=row, column=col).value = row * col

# 最后一定要保存！
wb.save(filename='write.xlsx')
