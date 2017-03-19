# encoding: utf-8
from openpyxl import Workbook  #Workbook 首字母大写
from openpyxl import load_workbook  #打开文件  首字母不用大写

wb = Workbook()  # 内存中创建一个excel文件

sheet1 = wb.get_active_sheet()
print (sheet1.title)

# 创建新的sheet
sheet2 = wb.create_sheet(title=u'新创建的sheet标题')

# 修改sheet1的title名字
sheet1.title = u'写入乘法表'

# 在sheet1中写入乘法表

for row in range(1, 10):
    for col in range(1, 10):
        sheet1.cell(row=row, column=col).value = row * col

# 保存文件

wb.save(filename=u'乘法表练习.xlsx')

# 练习读取excel文件

ws = load_workbook(filename='乘法表练习.xlsx')

sheet_names = ws.get_sheet_names()   # 获取所有sheet页的title
sheet_names_0 = sheet_names[1]

wss = ws.get_sheet_by_name(sheet_names[0])


rows = wss.rows
cols = wss.columns

# 通过坐标读取值
print(wss.cell('B8').value)  # B 表示列，8 表示行
print(wss['B8'].value)
print(wss.cell(row=9, column=2).value)


